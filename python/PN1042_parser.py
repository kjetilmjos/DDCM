#-*- coding: utf-8 -*-

import csv

from openpyxl import load_workbook
wb2 = load_workbook('tmppi_data-apr2016.xlsx')
ws1 = wb2.active

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK']

tags = []
description = []
unit = []
values = []

#---------------------Read columns and assign to arrays------------------------#

k = 3

while k <=  ws1.max_row:
    tags.append(ws1['A'+ str(k)].value)
    description.append(ws1['B'+ str(k)].value)
    unit.append(ws1['C'+ str(k)].value)
    k += 1
number_of_dates = ws1.max_column
s = 3
ks = 0
while s < number_of_dates:

    l = 1
    tmp = []
    while l <=  ws1.max_row:
        if (isinstance(ws1[columns[s]+ str(l)].value, str)):
            tmp.append(0)
        else:
            tmp.append(ws1[columns[s]+ str(l)].value)
        l += 1
    values.append(tmp)
    ks += 1
    s += 1

#---------------------/Read columns and assign to arrays------------------------#

with open('Parser_output.csv', 'w', newline='') as csvfile:
    file_write = csv.writer(csvfile, delimiter= ',')
    kk = 0
    ff = 0
    while kk < number_of_dates-3:
        ll = 2
        for i in range(len(tags)):
            data = [[tags[i], description[i], unit[i], values[ff][1],values[kk][ll] ]]
            file_write.writerows(data)
            ll += 1
        ff += 1
        kk += 1
