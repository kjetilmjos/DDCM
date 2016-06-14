from openpyxl import load_workbook
from dateutil.parser import parse
import datetime

columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK']

wb2 = load_workbook('./upload/excel_export.xlsx') # add a . to path start if not started from node server
ws1 = wb2.active

allowed_columns = 36

errors = 0


if ws1.max_column > allowed_columns:
    errors = 1
    print("Too many columns in use, is this only one month?")
s = 0

while s <= ws1.max_column-4:
    try_date = ws1[columns[s]+ str(2)].value
    if isinstance(try_date,datetime.datetime) == False:
        errors = 1
        print("Date not valid")

    s += 1

if errors ==1:
    print("Input file not valid")
else:
    print("File OK")
